"""Delivery layer: write the formatted Excel workbook."""

import os
from datetime import datetime

import pandas as pd

import config

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


def build_metadata(df):
    """Provenance for the workbook: what this is, where it came from, when."""
    rows = []

    for series_id, meta in config.SERIES.items():
        g = df[df["series_id"] == series_id]
        if g.empty:
            rows.append({
                "Series ID": series_id,
                "Description": meta["label"],
                "Units": meta["units"],
                "Kind": meta["kind"],
                "Observations": 0,
                "First Date": None,
                "Last Date": None,
                "Missing Values": None,
                "Note": "NO DATA RETURNED — this series is absent from the report",
            })
            continue

        rows.append({
            "Series ID": series_id,
            "Description": meta["label"],
            "Units": meta["units"],
            "Kind": meta["kind"],
            "Observations": len(g),
            "First Date": g["date"].min(),
            "Last Date": g["date"].max(),
            "Missing Values": int(g["value"].isna().sum()),
            "Note": "",
        })

    meta_df = pd.DataFrame(rows)
    meta_df.attrs["generated_at"] = datetime.now()
    return meta_df

def build_changes_table(df):
    """Full MoM/YoY history, presented rather than raw."""
    out = df[[
        "date", "label", "value", "mom_change",
        "yoy_change", "mom_change_3m_avg", "change_unit",
    ]].copy()

    out = out.rename(columns={
        "date": "Date",
        "label": "Series",
        "value": "Value",
        "mom_change": "MoM Change",
        "yoy_change": "YoY Change",
        "mom_change_3m_avg": "MoM Change (3M Avg)",
        "change_unit": "Unit",
    })

    return out.sort_values(["Series", "Date"]).reset_index(drop=True)

def write_excel(snapshot, wide, changes, metadata):
    """Write the four-sheet workbook. Returns the path."""
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    path = os.path.join(config.OUTPUT_DIR, config.EXCEL_NAME)

    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        snapshot.to_excel(writer, sheet_name="Summary", index=False)
        wide.to_excel(writer, sheet_name="Data")
        changes.to_excel(writer, sheet_name="Changes", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False, startrow=2)

        _style_workbook(writer, snapshot, wide, changes, metadata)

    return path

HEADER_FONT = Font(bold=True)
GREEN = Font(color="006100")
RED = Font(color="9C0006")


def _col_letter(ws, header_row, name):
    for cell in ws[header_row]:
        if cell.value == name:
            return get_column_letter(cell.column)
    return None


def _style_header(ws, header_row=1):
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, min_width=10, max_width=34):
    for col in ws.columns:
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def _format_column(ws, header_row, name, fmt):
    letter = _col_letter(ws, header_row, name)
    if letter is None:
        return
    for row in range(header_row + 1, ws.max_row + 1):
        ws[f"{letter}{row}"].number_format = fmt


def _color_changes(ws, header_row, name):
    letter = _col_letter(ws, header_row, name)
    if letter is None:
        return
    rng = f"{letter}{header_row + 1}:{letter}{ws.max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThan", formula=["0"], font=GREEN)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0"], font=RED)
    )


def _style_workbook(writer, snapshot, wide, changes, metadata):
    # ---- Summary ----
    ws = writer.sheets["Summary"]
    _style_header(ws)
    ws.freeze_panes = "A2"
    _format_column(ws, 1, "Value", "#,##0.00")
    for name in ("MoM", "YoY"):
        _format_column(ws, 1, name, "0.00")
        _color_changes(ws, 1, name)
    _autosize(ws)

    # ---- Data ----
    ws = writer.sheets["Data"]
    _style_header(ws)
    ws.freeze_panes = "B2"
    for col in range(2, ws.max_column + 1):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = "#,##0.00"
    _autosize(ws)

    # ---- Changes ----
    ws = writer.sheets["Changes"]
    _style_header(ws)
    ws.freeze_panes = "A2"
    _format_column(ws, 1, "Value", "#,##0.00")
    for name in ("MoM Change", "YoY Change", "MoM Change (3M Avg)"):
        _format_column(ws, 1, name, "0.00")
        _color_changes(ws, 1, name)
    _autosize(ws)

    # ---- Metadata ----
    ws = writer.sheets["Metadata"]
    ws["A1"] = f"Generated {metadata.attrs['generated_at']:%Y-%m-%d %H:%M} | Source: FRED (Federal Reserve Bank of St. Louis)"
    ws["A1"].font = HEADER_FONT
    _style_header(ws, header_row=3)
    _format_column(ws, 3, "Observations", "#,##0")
    _autosize(ws)